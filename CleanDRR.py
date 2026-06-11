import tkinter as tk
from tkinter import ttk, messagebox
import threading
import os
import io
import pandas as pd
from openpyxl.styles import PatternFill, Font
import openpyxl

# ── CONFIG ────────────────────────────────────────────────────────────────────
STATUS_COL_NAME    = "Status"
STATUS_COL_INDEX   = 9
DATE_COL_NAME      = "Date"
TIME_COL_NAME      = "Time"
COL_E_INDEX        = 4          # Account Number column
DIALED_COL_INDEX   = 5          # Column F = Dialed Number (index 5)
PTP_AMOUNT_COL     = "PTP Amount"
CLAIM_PAID_COL     = "Claim Paid Amount"
REMARK_COL         = "Remark"
REMOVE_STATUSES    = {"BP", "REACTIVE", "SMS FAILED", "NEW", "ABORTED"}
COLS_TO_DROP_START = 27
COLS_TO_DROP_END   = 50

# ── COLORS ────────────────────────────────────────────────────────────────────
BG        = "#0F1117"
CARD      = "#1A1D27"
BORDER    = "#2A2D3E"
ACCENT    = "#4F8EF7"
ACCENT2   = "#7C3AED"
SUCCESS   = "#22C55E"
DANGER    = "#EF4444"
WARNING   = "#F59E0B"
TEXT      = "#F1F5F9"
SUBTEXT   = "#94A3B8"
HEADER_BG = "#1E2235"

# ── PROCESSING LOGIC ─────────────────────────────────────────────────────────
def process_file(filepath):
    df_headers = pd.read_excel(filepath, nrows=0)

    # Force Column E (Account Number) as string
    str_cols = {}
    if COL_E_INDEX < len(df_headers.columns):
        str_cols[df_headers.columns[COL_E_INDEX]] = str

    col_e_name      = df_headers.columns[COL_E_INDEX]      if COL_E_INDEX      < len(df_headers.columns) else None
    dialed_col_name = df_headers.columns[DIALED_COL_INDEX] if DIALED_COL_INDEX < len(df_headers.columns) else None

    # Also force Dialed Number column as string
    if dialed_col_name:
        str_cols[dialed_col_name] = str

    # Try to find "Dialed Number" by name if index is wrong
    for col in df_headers.columns:
        if "dialed" in str(col).lower():
            str_cols[col] = str
            dialed_col_name = col
            break

    df = pd.read_excel(filepath, dtype=str_cols)

    def strip_decimal(x):
        """Remove decimal portion from numeric strings — e.g. 639208174621.0 -> 639208174621."""
        if pd.isna(x) or str(x).strip() in ("", "nan"):
            return ""
        s = str(x).strip()
        if "." in s:
            integer_part = s.split(".")[0]
            if integer_part.lstrip("-").isdigit():
                return integer_part
        return s

    def clean_account(x):
        """Strip .0 only if safe — preserve leading zeros."""
        if pd.isna(x) or str(x).strip() in ("", "nan"):
            return ""
        s = str(x).strip()
        if s.endswith(".0") and s[:-2].lstrip("-").isdigit() and not s[:-2].startswith("0"):
            return s[:-2]
        return s

    if col_e_name and col_e_name in df.columns:
        df[col_e_name] = df[col_e_name].apply(clean_account)

    # Fix Dialed Number — strip decimals from ALL rows
    if dialed_col_name and dialed_col_name in df.columns:
        df[dialed_col_name] = df[dialed_col_name].apply(strip_decimal)
    else:
        # Fallback: fix any column whose name contains "dialed"
        for col in df.columns:
            if "dialed" in str(col).lower():
                df[col] = df[col].apply(strip_decimal)
                dialed_col_name = col
                break

    STATUS_COL = STATUS_COL_NAME if STATUS_COL_NAME in df.columns else (
        df.columns[STATUS_COL_INDEX] if STATUS_COL_INDEX < len(df.columns) else None
    )
    if not STATUS_COL:
        raise ValueError("Status column not found.")

    status_normalized = df[STATUS_COL].astype(str).str.strip().str.upper()
    is_blank          = df[STATUS_COL].isna() | (df[STATUS_COL].astype(str).str.strip() == "")
    has_cease         = status_normalized.str.contains("CEASE", na=False)
    is_removable      = status_normalized.isin(REMOVE_STATUSES) | is_blank | has_cease

    if PTP_AMOUNT_COL in df.columns:
        has_ptp       = status_normalized.str.contains(r"\bPTP\b", regex=True, na=False)
        ptp_numeric   = pd.to_numeric(df[PTP_AMOUNT_COL].astype(str).str.replace(",", "", regex=False), errors="coerce").fillna(0)
        ptp_has_value = has_ptp & (ptp_numeric > 0)
        ptp_no_value  = has_ptp & (ptp_numeric <= 0)
        is_removable  = (is_removable | ptp_no_value) & ~ptp_has_value

    if CLAIM_PAID_COL in df.columns:
        has_kept       = status_normalized.str.contains(r"\bKEPT\b", regex=True, na=False)
        claim_numeric  = pd.to_numeric(df[CLAIM_PAID_COL].astype(str).str.replace(",", "", regex=False), errors="coerce").fillna(0)
        kept_has_value = has_kept & (claim_numeric > 0)
        kept_no_value  = has_kept & (claim_numeric <= 0)
        is_removable   = (is_removable | kept_no_value) & ~kept_has_value

    cleaned_df = df[~is_removable].copy()
    removed_df = df[is_removable].copy()

    # ── ADD REMOVED REASON COLUMN ─────────────────────────────────────────────
    def get_reason(row):
        s = str(row[STATUS_COL]).strip().upper()
        if s in ("", "NAN"):
            return "Blank Status"
        if s in REMOVE_STATUSES:
            return f"Status: {str(row[STATUS_COL]).strip()}"
        if "CEASE" in s:
            return f"Status contains CEASE: {str(row[STATUS_COL]).strip()}"
        if "PTP" in s:
            return "PTP with no PTP Amount"
        if "KEPT" in s:
            return "KEPT with no Claim Paid Amount"
        return f"Status: {str(row[STATUS_COL]).strip()}"

    removed_df.insert(0, "Removed Reason", removed_df.apply(get_reason, axis=1))

    # ── REMARK: Action: PTP → Action: SRP if status not PTP ─────────────────
    srp_mask   = pd.Series([False] * len(cleaned_df), index=cleaned_df.index)
    remarks_df = pd.DataFrame()   # for the Remarks Changes tab

    if REMARK_COL in cleaned_df.columns:
        remark_norm    = cleaned_df[REMARK_COL].astype(str).str.strip().str.upper()
        status_clean   = cleaned_df[STATUS_COL].astype(str).str.strip().str.upper()
        has_action_ptp = remark_norm.str.contains(r"ACTION\s*:\s*PTP", regex=True, na=False)
        # Do NOT change if status contains PTP or KEPT (including KEPT_FULL, KEPT_PARTIAL etc.)
        status_not_ptp_kept = (
            ~status_clean.str.contains("PTP",  na=False) &
            ~status_clean.str.contains("KEPT", na=False)
        )
        srp_mask = has_action_ptp & status_not_ptp_kept

        # Build remarks change log BEFORE applying the change
        if srp_mask.any():
            changed_rows = cleaned_df.loc[srp_mask].copy()
            remarks_df = pd.DataFrame({
                "Row #":        range(1, srp_mask.sum() + 1),
                STATUS_COL:     changed_rows[STATUS_COL].values,
                REMARK_COL + " (Before)": changed_rows[REMARK_COL].astype(str).values,
                REMARK_COL + " (After)":  changed_rows[REMARK_COL].astype(str).str.replace(
                    r"(?i)Action\s*:\s*PTP", "Action: SRP", regex=True).values,
            })

        cleaned_df.loc[srp_mask, REMARK_COL] = cleaned_df.loc[srp_mask, REMARK_COL].astype(str).str.replace(
            r"(?i)Action\s*:\s*PTP", "Action: SRP", regex=True)

    # ── DROP COLUMNS AB–AY ────────────────────────────────────────────────────
    cols_to_drop = [df.columns[i] for i in range(COLS_TO_DROP_START, min(COLS_TO_DROP_END + 1, len(df.columns)))]
    cleaned_df.drop(columns=cols_to_drop, inplace=True, errors="ignore")
    removed_df.drop(columns=cols_to_drop, inplace=True, errors="ignore")

    # ── DATE / TIME ───────────────────────────────────────────────────────────
    if DATE_COL_NAME in cleaned_df.columns:
        cleaned_df[DATE_COL_NAME] = pd.to_datetime(cleaned_df[DATE_COL_NAME], errors="coerce").dt.strftime("%m-%d-%Y")

    if DATE_COL_NAME in cleaned_df.columns and TIME_COL_NAME in cleaned_df.columns:
        date_str = pd.to_datetime(cleaned_df[DATE_COL_NAME], errors="coerce").dt.strftime("%m-%d-%Y")
        time_str = pd.to_datetime(cleaned_df[TIME_COL_NAME], errors="coerce").dt.strftime("%I:%M:%S %p")
        cleaned_df[TIME_COL_NAME] = date_str + " " + time_str

    stats = {
        "total":       len(df),
        "retained":    len(cleaned_df),
        "removed":     len(removed_df),
        "srp_changed": int(srp_mask.sum()),
    }

    # ── EXPORT ────────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for sheet_name, frame in [("Cleaned", cleaned_df), ("Removed", removed_df)]:
            frame.to_excel(writer, index=False, sheet_name=sheet_name)
            ws = writer.sheets[sheet_name]
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                # Highlight "Removed Reason" header in orange
                if cell.value == "Removed Reason":
                    cell.fill = PatternFill("solid", fgColor="C05621")
                else:
                    cell.fill = PatternFill("solid", fgColor="1E2235")
            # Highlight Removed Reason cells in light orange
            if "Removed Reason" in frame.columns:
                reason_idx = list(frame.columns).index("Removed Reason") + 1
                for row in ws.iter_rows(min_row=2, min_col=reason_idx, max_col=reason_idx):
                    for cell in row:
                        cell.fill = PatternFill("solid", fgColor="FFF3E0")
                        cell.font = Font(bold=True, color="C05621")
            # Force dialed number and account number cols as text in Excel
            for col_name in [col_e_name, dialed_col_name]:
                if col_name and col_name in frame.columns:
                    col_idx = list(frame.columns).index(col_name) + 1
                    for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                        for cell in row:
                            cell.number_format = "@"
            for col in ws.columns:
                max_len = max((len(str(c.value)) for c in col if c.value), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

        if not remarks_df.empty:
            remarks_df.to_excel(writer, index=False, sheet_name="Remarks Changes")
            ws = writer.sheets["Remarks Changes"]
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="7C3AED")
            for col in ws.columns:
                max_len = max((len(str(c.value)) for c in col if c.value), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    buf.seek(0)
    return cleaned_df, removed_df, remarks_df, stats, buf.read(), col_e_name, dialed_col_name


# ── TOAST NOTIFICATION ────────────────────────────────────────────────────────
def show_toast(parent, message, color=SUCCESS, duration=3000):
    toast = tk.Toplevel(parent)
    toast.overrideredirect(True)
    toast.attributes("-topmost", True)
    toast.attributes("-alpha", 0.95)

    frame = tk.Frame(toast, bg=color, padx=20, pady=12)
    frame.pack()
    tk.Label(frame, text="✔  " + message, font=("Segoe UI", 10, "bold"),
             bg=color, fg="white").pack()

    # Position bottom-right of parent
    parent.update_idletasks()
    pw = parent.winfo_width()
    ph = parent.winfo_height()
    px = parent.winfo_x()
    py = parent.winfo_y()
    toast.update_idletasks()
    tw = toast.winfo_reqwidth()
    th = toast.winfo_reqheight()
    toast.geometry(f"+{px + pw - tw - 20}+{py + ph - th - 40}")

    # Fade out then destroy
    def fade_out(alpha=0.95):
        alpha -= 0.05
        if alpha <= 0:
            toast.destroy()
        else:
            toast.attributes("-alpha", alpha)
            toast.after(40, lambda: fade_out(alpha))

    toast.after(duration, fade_out)


# ── APP ───────────────────────────────────────────────────────────────────────
class CleanerApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Excel Data Cleaner")
        self.geometry("1140x740")
        self.minsize(900, 600)
        self.configure(bg=BG)
        self.resizable(True, True)

        self.file_path          = None
        self.output_bytes       = None
        self.cleaned_df         = None
        self.removed_df         = None
        self.removed_reason_df  = None
        self.remarks_df         = None

        self._build_ui()

    def _build_ui(self):
        # ── Header ────────────────────────────────────────────────────────────
        header = tk.Frame(self, bg=HEADER_BG, height=56)
        header.pack(fill="x")
        header.pack_propagate(False)
        tk.Label(header, text="⬡  Excel Data Cleaner", font=("Segoe UI", 14, "bold"),
                 bg=HEADER_BG, fg=TEXT).pack(side="left", padx=20, pady=14)
        tk.Label(header, text="Drop · Preview · Verify · Download",
                 font=("Segoe UI", 9), bg=HEADER_BG, fg=SUBTEXT).pack(side="left", padx=4)

        # ── Body ──────────────────────────────────────────────────────────────
        body = tk.Frame(self, bg=BG)
        body.pack(fill="both", expand=True, padx=20, pady=16)

        # ── LEFT PANEL ────────────────────────────────────────────────────────
        left = tk.Frame(body, bg=BG, width=300)
        left.pack(side="left", fill="y", padx=(0, 14))
        left.pack_propagate(False)

        # Drop zone
        drop_frame = tk.Frame(left, bg=CARD, highlightthickness=2, highlightbackground=BORDER)
        drop_frame.pack(fill="x", pady=(0, 12))
        drop_inner = tk.Frame(drop_frame, bg=CARD)
        drop_inner.pack(fill="both", padx=20, pady=24)
        tk.Label(drop_inner, text="📂", font=("Segoe UI", 26), bg=CARD, fg=ACCENT).pack()
        tk.Label(drop_inner, text="Drop Excel file here", font=("Segoe UI", 11, "bold"),
                 bg=CARD, fg=TEXT).pack(pady=(6, 2))
        tk.Label(drop_inner, text=".xlsx  ·  .xlsm  ·  .xls", font=("Segoe UI", 8),
                 bg=CARD, fg=SUBTEXT).pack()
        try:
            drop_frame.drop_target_register('DND_Files')
            drop_frame.dnd_bind('<<Drop>>', self._on_drop)
            tk.Label(drop_inner, text="or", font=("Segoe UI", 8), bg=CARD, fg=SUBTEXT).pack(pady=(8, 4))
        except Exception:
            pass
        tk.Button(drop_inner, text="Browse File", font=("Segoe UI", 9, "bold"),
                  bg=ACCENT, fg="white", relief="flat", padx=16, pady=6,
                  cursor="hand2", command=self._browse).pack()

        self.file_label = tk.Label(left, text="No file selected", font=("Segoe UI", 8),
                                   bg=BG, fg=SUBTEXT, wraplength=270, justify="left")
        self.file_label.pack(fill="x", pady=(0, 12))

        # Stats
        tk.Label(left, text="SUMMARY", font=("Segoe UI", 8, "bold"),
                 bg=BG, fg=SUBTEXT).pack(anchor="w", pady=(0, 6))

        self.stat_vars = {}
        for key, label, color, icon in [
            ("total",    "Total Rows",     TEXT,    "📋"),
            ("retained", "Rows Retained",  SUCCESS, "✅"),
            ("removed",  "Rows Removed",   DANGER,  "🗑"),
            ("srp",      "Remarks Fixed",  WARNING, "✏️"),
        ]:
            card = tk.Frame(left, bg=CARD, highlightthickness=1, highlightbackground=BORDER)
            card.pack(fill="x", pady=3)
            inner = tk.Frame(card, bg=CARD)
            inner.pack(fill="x", padx=12, pady=8)
            tk.Label(inner, text=icon, font=("Segoe UI", 11), bg=CARD, fg=color).pack(side="left")
            tk.Label(inner, text=label, font=("Segoe UI", 9), bg=CARD, fg=SUBTEXT).pack(side="left", padx=8)
            var = tk.StringVar(value="—")
            self.stat_vars[key] = var
            tk.Label(inner, textvariable=var, font=("Segoe UI", 11, "bold"),
                     bg=CARD, fg=color).pack(side="right")

        # Buttons
        self.process_btn = tk.Button(left, text="▶  Run Cleaner",
                                     font=("Segoe UI", 10, "bold"), bg=ACCENT, fg="white",
                                     relief="flat", pady=10, cursor="hand2",
                                     state="disabled", command=self._run_process)
        self.process_btn.pack(fill="x", pady=(14, 6))

        self.download_btn = tk.Button(left, text="⬇  Download Output",
                                      font=("Segoe UI", 9, "bold"), bg=SUCCESS, fg="white",
                                      relief="flat", pady=10, cursor="hand2",
                                      state="disabled", command=self._download)
        self.download_btn.pack(fill="x")

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("TProgressbar", troughcolor=CARD, background=ACCENT, thickness=4)
        self.progress = ttk.Progressbar(left, mode="indeterminate")
        self.progress.pack(fill="x", pady=(10, 0))

        # ── RIGHT PANEL ───────────────────────────────────────────────────────
        right = tk.Frame(body, bg=BG)
        right.pack(side="left", fill="both", expand=True)

        # Tab bar
        tab_bar = tk.Frame(right, bg=BG)
        tab_bar.pack(fill="x", pady=(0, 8))

        self.active_tab = tk.StringVar(value="cleaned")
        self.tab_btns   = {}
        for tab_id, tab_label, tab_color in [
            ("cleaned",         "✅ Cleaned Rows",      ACCENT),
            ("removed",         "🗑 Removed Rows",      CARD),
            ("removed_reason",  "📋 Removed Status",    CARD),
            ("remarks",         "✏️ Remarks Changes",   CARD),
        ]:
            btn = tk.Button(tab_bar, text=tab_label,
                            font=("Segoe UI", 9, "bold"),
                            bg=tab_color, fg="white", relief="flat",
                            padx=14, pady=7, cursor="hand2",
                            command=lambda t=tab_id: self._switch_tab(t))
            btn.pack(side="left", padx=(0, 6))
            self.tab_btns[tab_id] = btn

        # Search
        search_frame = tk.Frame(right, bg=CARD, highlightthickness=1, highlightbackground=BORDER)
        search_frame.pack(fill="x", pady=(0, 8))
        tk.Label(search_frame, text="🔍", bg=CARD, fg=SUBTEXT, font=("Segoe UI", 10)).pack(side="left", padx=8)
        self.search_var = tk.StringVar()
        self.search_var.trace_add("write", self._on_search)
        tk.Entry(search_frame, textvariable=self.search_var, font=("Segoe UI", 9),
                 bg=CARD, fg=TEXT, insertbackground=TEXT, relief="flat", bd=0
                 ).pack(side="left", fill="x", expand=True, pady=8, padx=4)
        tk.Label(search_frame, text="Filter rows…", bg=CARD, fg=SUBTEXT,
                 font=("Segoe UI", 8)).pack(side="right", padx=8)

        # Table
        table_frame = tk.Frame(right, bg=BG)
        table_frame.pack(fill="both", expand=True)
        vsb = ttk.Scrollbar(table_frame, orient="vertical")
        hsb = ttk.Scrollbar(table_frame, orient="horizontal")
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")

        style.configure("Custom.Treeview", background=CARD, foreground=TEXT,
                         fieldbackground=CARD, rowheight=28, font=("Segoe UI", 9), borderwidth=0)
        style.configure("Custom.Treeview.Heading", background=HEADER_BG, foreground=TEXT,
                         font=("Segoe UI", 9, "bold"), relief="flat")
        style.map("Custom.Treeview",
                  background=[("selected", ACCENT2)], foreground=[("selected", "white")])

        self.tree = ttk.Treeview(table_frame, style="Custom.Treeview",
                                  yscrollcommand=vsb.set, xscrollcommand=hsb.set,
                                  show="headings", selectmode="browse")
        self.tree.pack(fill="both", expand=True)
        vsb.config(command=self.tree.yview)
        hsb.config(command=self.tree.xview)

        self.row_count_label = tk.Label(right, text="", font=("Segoe UI", 8), bg=BG, fg=SUBTEXT)
        self.row_count_label.pack(anchor="e", pady=(4, 0))

        # ── Watermark footer ──────────────────────────────────────────────────
        watermark = tk.Frame(self, bg="#0A0C14")
        watermark.pack(fill="x", side="bottom")

        left_mark = tk.Frame(watermark, bg="#0A0C14")
        left_mark.pack(side="left", padx=16, pady=6)
        tk.Label(left_mark, text="Created by  Vincent Corocoto",
                 font=("Segoe UI", 8, "bold"), bg="#0A0C14", fg=ACCENT).pack(side="left")
        tk.Label(left_mark, text="  ·  09567796275",
                 font=("Segoe UI", 8), bg="#0A0C14", fg=SUBTEXT).pack(side="left")

        tk.Label(watermark,
                 text='"Kapag ang palay naging bigas, May bumayo."',
                 font=("Segoe UI", 8, "italic"), bg="#0A0C14", fg="#4A5568"
                 ).pack(side="right", padx=16, pady=6)

        # Status bar
        self.status_var = tk.StringVar(value="Ready — drop or browse an Excel file to begin.")
        tk.Label(self, textvariable=self.status_var, font=("Segoe UI", 8),
                 bg=HEADER_BG, fg=SUBTEXT, anchor="w", padx=16, pady=6).pack(fill="x", side="bottom")

    # ── FILE LOAD ─────────────────────────────────────────────────────────────
    def _on_drop(self, event):
        self._load_file(event.data.strip().strip("{}"))

    def _browse(self):
        from tkinter import filedialog
        path = filedialog.askopenfilename(title="Select Excel file",
                                          filetypes=[("Excel files", "*.xlsx *.xlsm *.xls")])
        if path:
            self._load_file(path)

    def _load_file(self, path):
        if not os.path.exists(path):
            messagebox.showerror("File not found", f"Cannot find:\n{path}")
            return
        self.file_path = path
        self.file_label.config(text=f"📄 {os.path.basename(path)}", fg=ACCENT)
        self.process_btn.config(state="normal")
        self.download_btn.config(state="disabled")
        self.status_var.set(f"Loaded: {os.path.basename(path)}  —  Click ▶ Run Cleaner to process.")
        self._clear_table()
        for k in self.stat_vars:
            self.stat_vars[k].set("—")

    # ── PROCESSING ────────────────────────────────────────────────────────────
    def _run_process(self):
        if not self.file_path:
            return
        self.process_btn.config(state="disabled")
        self.download_btn.config(state="disabled")
        self.status_var.set("Processing…")
        self.progress.start(10)
        threading.Thread(target=self._process_thread, daemon=True).start()

    def _process_thread(self):
        try:
            result = process_file(self.file_path)
            self.after(0, lambda: self._on_done(*result))
        except Exception as e:
            self.after(0, lambda: self._on_error(str(e)))

    def _on_done(self, cleaned, removed, remarks, stats, out_bytes, col_e, dialed_col):
        self.progress.stop()
        self.cleaned_df         = cleaned
        self.removed_df         = removed
        # Removed Status tab: only show rows matching exact REMOVE_STATUSES
        if "Removed Reason" in removed.columns:
            status_col = [c for c in removed.columns if c != "Removed Reason"]
            mask = removed["Removed Reason"].str.startswith("Status: ")
            self.removed_reason_df = removed[mask][["Removed Reason"] + status_col].copy()
        else:
            self.removed_reason_df = removed.copy()
        self.remarks_df         = remarks
        self.output_bytes       = out_bytes

        self.stat_vars["total"].set(str(stats["total"]))
        self.stat_vars["retained"].set(str(stats["retained"]))
        self.stat_vars["removed"].set(str(stats["removed"]))
        self.stat_vars["srp"].set(str(stats["srp_changed"]))

        self.download_btn.config(state="normal")
        self.process_btn.config(state="normal")
        self.status_var.set(
            f"Done  ·  {stats['retained']} retained  ·  {stats['removed']} removed  ·  "
            f"{stats['srp_changed']} remarks → SRP"
        )
        self._switch_tab("cleaned")

    def _on_error(self, msg):
        self.progress.stop()
        self.process_btn.config(state="normal")
        self.status_var.set(f"Error: {msg}")
        messagebox.showerror("Processing Error", msg)

    # ── TABS & TABLE ──────────────────────────────────────────────────────────
    def _switch_tab(self, tab_id):
        self.active_tab.set(tab_id)
        colors = {"cleaned": CARD, "removed": CARD, "removed_reason": CARD, "remarks": CARD}
        colors[tab_id] = ACCENT2 if tab_id in ("remarks", "removed_reason") else ACCENT
        for tid, btn in self.tab_btns.items():
            btn.config(bg=colors[tid])

        df_map = {
            "cleaned":        self.cleaned_df,
            "removed":        self.removed_df,
            "removed_reason": self.removed_reason_df,
            "remarks":        self.remarks_df,
        }
        df = df_map.get(tab_id)

        empty_msgs = {
            "remarks":        "No remark changes found.",
            "removed_reason": "No removed rows found.",
            "removed":        "No removed rows.",
            "cleaned":        "No retained rows.",
        }

        if df is not None and not (hasattr(df, "empty") and df.empty):
            self._populate_table(df)
        else:
            self._clear_table()
            self.row_count_label.config(text=empty_msgs.get(tab_id, "No rows."))

    def _populate_table(self, df, filter_text=""):
        self._clear_table()
        if df is None or df.empty:
            self.row_count_label.config(text="No rows to display.")
            return

        # Show ALL columns — no truncation
        cols = list(df.columns)
        self.tree["columns"] = cols
        for col in cols:
            # Auto-size column width based on header + sample data
            max_len = max(
                len(str(col)),
                df[col].astype(str).str.len().quantile(0.90) if len(df) > 0 else 10
            )
            col_width = max(100, min(int(max_len) * 8, 280))
            self.tree.heading(col, text=col, anchor="w",
                              command=lambda c=col: self._sort_col(c))
            self.tree.column(col, width=col_width, minwidth=60, anchor="w", stretch=False)

        filt  = filter_text.lower()
        count = 0
        for _, row in df.iterrows():
            vals = [str(v) if pd.notna(v) else "" for v in row[cols]]
            if filt and not any(filt in v.lower() for v in vals):
                continue
            tag = "even" if count % 2 == 0 else "odd"
            self.tree.insert("", "end", values=vals, tags=(tag,))
            count += 1

        self.tree.tag_configure("even", background=CARD)
        self.tree.tag_configure("odd",  background="#14172A")
        self.row_count_label.config(
            text=f"Showing {count} of {len(df)} rows  ·  {len(df.columns)} columns  —  scroll horizontally to see all"
        )

    def _sort_col(self, col):
        """Sort table by column when header is clicked."""
        tab = self.active_tab.get()
        df  = {"cleaned": self.cleaned_df, "removed": self.removed_df, "remarks": self.remarks_df}.get(tab)
        if df is None:
            return
        asc = getattr(self, "_sort_asc", {})
        ascending = not asc.get(col, True)
        asc[col] = ascending
        self._sort_asc = asc
        try:
            df_sorted = df.sort_values(by=col, ascending=ascending, key=lambda x: x.astype(str).str.lower())
        except Exception:
            df_sorted = df
        if tab == "cleaned":
            self.cleaned_df = df_sorted
        elif tab == "removed":
            self.removed_df = df_sorted
        else:
            self.remarks_df = df_sorted
        self._populate_table(df_sorted, self.search_var.get())

    def _clear_table(self):
        self.tree.delete(*self.tree.get_children())
        self.tree["columns"] = []
        self.row_count_label.config(text="")

    def _on_search(self, *_):
        tab = self.active_tab.get()
        df  = {
            "cleaned":        self.cleaned_df,
            "removed":        self.removed_df,
            "removed_reason": self.removed_reason_df,
            "remarks":        self.remarks_df,
        }.get(tab)
        if df is not None:
            self._populate_table(df, self.search_var.get())

    # ── DOWNLOAD ──────────────────────────────────────────────────────────────
    def _download(self):
        from tkinter import filedialog
        save_path = filedialog.asksaveasfilename(
            title="Save output as",
            defaultextension=".xlsx",
            initialfile="cleaned_output.xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if not save_path:
            return
        try:
            with open(save_path, "wb") as f:
                f.write(self.output_bytes)
            self.status_var.set(f"Saved → {save_path}")
            show_toast(self, f"Saved: {os.path.basename(save_path)}", color=SUCCESS)
        except Exception as e:
            show_toast(self, f"Save failed: {e}", color=DANGER, duration=5000)


if __name__ == "__main__":
    app = CleanerApp()
    app.mainloop()