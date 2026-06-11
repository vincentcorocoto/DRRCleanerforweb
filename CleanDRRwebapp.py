import streamlit as st
import pandas as pd
import io
import json
from openpyxl.styles import PatternFill, Font
import openpyxl
import streamlit.components.v1 as components

# ── PAGE CONFIG ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Excel Data Cleaner",
    page_icon="⬡",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── CONFIG ────────────────────────────────────────────────────────────────────
STATUS_COL_NAME    = "Status"
STATUS_COL_INDEX   = 9
DATE_COL_NAME      = "Date"
TIME_COL_NAME      = "Time"
COL_E_INDEX        = 4
DIALED_COL_INDEX   = 5
PTP_AMOUNT_COL     = "PTP Amount"
CLAIM_PAID_COL     = "Claim Paid Amount"
REMARK_COL         = "Remark"
REMOVE_STATUSES    = {"BP", "REACTIVE", "SMS FAILED", "NEW", "ABORTED"}
COLS_TO_DROP_START = 27
COLS_TO_DROP_END   = 50

# ── GLOBAL CSS ────────────────────────────────────────────────────────────────
st.markdown("""
<style>
  /* ── Base ── */
  [data-testid="stAppViewContainer"] { background: #0F1117; }
  [data-testid="stSidebar"]          { background: #1A1D27; border-right: 1px solid #2A2D3E; }
  [data-testid="stHeader"]           { background: #1E2235; }
  .block-container                   { padding-top: 1.5rem; padding-bottom: 1rem; }

  /* ── Sidebar labels ── */
  [data-testid="stSidebar"] label,
  [data-testid="stSidebar"] .stMarkdown p { color: #94A3B8 !important; font-size: 13px; }

  /* ── Metric cards ── */
  [data-testid="stMetric"]            { background: #1A1D27; border: 1px solid #2A2D3E;
                                        border-radius: 10px; padding: 12px 16px; }
  [data-testid="stMetricLabel"]       { color: #94A3B8 !important; font-size: 12px !important; }
  [data-testid="stMetricValue"]       { color: #F1F5F9 !important; font-size: 22px !important; }

  /* ── Tab bar ── */
  .stTabs [data-baseweb="tab-list"]   { background: #1A1D27; border-radius: 10px;
                                        padding: 4px; gap: 4px; border: 1px solid #2A2D3E; }
  .stTabs [data-baseweb="tab"]        { background: transparent; color: #94A3B8;
                                        border-radius: 7px; font-size: 13px; font-weight: 600;
                                        padding: 6px 16px; border: none; }
  .stTabs [aria-selected="true"]      { background: #4F8EF7 !important; color: white !important; }
  .stTabs [data-baseweb="tab-panel"]  { padding-top: 12px; }

  /* ── Buttons ── */
  .stButton > button {
    background: #4F8EF7; color: white; border: none;
    border-radius: 8px; font-weight: 700; font-size: 14px;
    padding: 10px 20px; width: 100%; cursor: pointer;
    transition: background 0.15s;
  }
  .stButton > button:hover { background: #3a7de8; }

  /* ── Download button ── */
  [data-testid="stDownloadButton"] > button {
    background: #22C55E !important; color: white !important; border: none !important;
    border-radius: 8px !important; font-weight: 700 !important; font-size: 14px !important;
    padding: 10px 20px !important; width: 100% !important;
  }
  [data-testid="stDownloadButton"] > button:hover { background: #16a34a !important; }

  /* ── File uploader ── */
  [data-testid="stFileUploader"] {
    background: #1A1D27; border: 2px dashed #2A2D3E;
    border-radius: 12px; padding: 8px;
  }
  [data-testid="stFileUploader"]:hover { border-color: #4F8EF7; }

  /* ── Search input ── */
  .stTextInput > div > div > input {
    background: #1A1D27 !important; color: #F1F5F9 !important;
    border: 1px solid #2A2D3E !important; border-radius: 8px !important;
    font-size: 13px !important;
  }

  /* ── Inspector panel ── */
  .inspector-panel {
    background: #12162B;
    border: 1px solid #4F8EF7;
    border-radius: 10px;
    padding: 12px 16px;
    margin-bottom: 10px;
  }
  .inspector-col  { color: #4F8EF7; font-size: 12px; font-weight: 700;
                    text-transform: uppercase; letter-spacing: 0.06em; margin-bottom: 6px; }
  .inspector-val  { color: #F1F5F9; font-size: 14px; line-height: 1.6;
                    word-break: break-all; white-space: pre-wrap; }
  .inspector-empty { color: #94A3B8; font-style: italic; }

  /* ── Status bar ── */
  .status-bar {
    background: #1E2235; border-radius: 8px; padding: 8px 14px;
    color: #94A3B8; font-size: 12px; margin-bottom: 10px;
  }

  /* ── Stat pill ── */
  .stat-pill {
    display: inline-block; padding: 3px 10px; border-radius: 99px;
    font-size: 12px; font-weight: 700;
  }

  /* ── Section header ── */
  .section-header {
    color: #94A3B8; font-size: 11px; font-weight: 700;
    text-transform: uppercase; letter-spacing: 0.08em;
    margin: 16px 0 8px;
  }

  /* ── Scrollbar ── */
  ::-webkit-scrollbar       { width: 6px; height: 6px; }
  ::-webkit-scrollbar-track { background: #1A1D27; }
  ::-webkit-scrollbar-thumb { background: #2A2D3E; border-radius: 3px; }

  /* ── Watermark ── */
  .watermark {
    text-align: center; color: #4A5568; font-size: 11px;
    padding: 12px 0; border-top: 1px solid #2A2D3E; margin-top: 16px;
  }
  .watermark span { color: #4F8EF7; font-weight: 700; }
</style>
""", unsafe_allow_html=True)


# ── PROCESSING LOGIC ─────────────────────────────────────────────────────────
def process_file(file_bytes):
    df_headers = pd.read_excel(io.BytesIO(file_bytes), nrows=0)

    str_cols = {}
    if COL_E_INDEX < len(df_headers.columns):
        str_cols[df_headers.columns[COL_E_INDEX]] = str

    col_e_name      = df_headers.columns[COL_E_INDEX]      if COL_E_INDEX      < len(df_headers.columns) else None
    dialed_col_name = df_headers.columns[DIALED_COL_INDEX] if DIALED_COL_INDEX < len(df_headers.columns) else None

    if dialed_col_name:
        str_cols[dialed_col_name] = str

    for col in df_headers.columns:
        if "dialed" in str(col).lower():
            str_cols[col] = str
            dialed_col_name = col
            break

    df = pd.read_excel(io.BytesIO(file_bytes), dtype=str_cols)

    def vectorized_strip_decimal(series):
        s          = series.fillna("").astype(str).str.strip()
        has_dot    = s.str.contains(".", regex=False)
        int_part   = s.str.split(".").str[0]
        is_numeric = int_part.str.lstrip("-").str.isdigit()
        return s.where(~(has_dot & is_numeric), int_part).replace("nan", "")

    def vectorized_clean_account(series):
        s         = series.fillna("").astype(str).str.strip()
        ends_dot0 = s.str.endswith(".0")
        base      = s.str[:-2]
        is_safe   = ends_dot0 & base.str.lstrip("-").str.isdigit() & ~base.str.startswith("0")
        return s.where(~is_safe, base).replace("nan", "")

    if col_e_name and col_e_name in df.columns:
        df[col_e_name] = vectorized_clean_account(df[col_e_name])

    if dialed_col_name and dialed_col_name in df.columns:
        df[dialed_col_name] = vectorized_strip_decimal(df[dialed_col_name])
    else:
        for col in df.columns:
            if "dialed" in str(col).lower():
                df[col] = vectorized_strip_decimal(df[col])
                dialed_col_name = col
                break

    STATUS_COL = STATUS_COL_NAME if STATUS_COL_NAME in df.columns else (
        df.columns[STATUS_COL_INDEX] if STATUS_COL_INDEX < len(df.columns) else None
    )
    if not STATUS_COL:
        raise ValueError("Status column not found.")

    status_normalized = df[STATUS_COL].astype(str).str.strip().str.upper()
    is_blank          = df[STATUS_COL].isna() | (df[STATUS_COL].astype(str).str.strip() == "")
    has_cease         = status_normalized.str.contains("CEASE", na=False)
    is_removable      = status_normalized.isin(REMOVE_STATUSES) | is_blank | has_cease

    if PTP_AMOUNT_COL in df.columns:
        has_ptp       = status_normalized.str.contains(r"\bPTP\b", regex=True, na=False)
        ptp_numeric   = pd.to_numeric(df[PTP_AMOUNT_COL].astype(str).str.replace(",", "", regex=False), errors="coerce").fillna(0)
        ptp_has_value = has_ptp & (ptp_numeric > 0)
        ptp_no_value  = has_ptp & (ptp_numeric <= 0)
        is_removable  = (is_removable | ptp_no_value) & ~ptp_has_value

    if CLAIM_PAID_COL in df.columns:
        has_kept       = status_normalized.str.contains(r"\bKEPT\b", regex=True, na=False)
        claim_numeric  = pd.to_numeric(df[CLAIM_PAID_COL].astype(str).str.replace(",", "", regex=False), errors="coerce").fillna(0)
        kept_has_value = has_kept & (claim_numeric > 0)
        kept_no_value  = has_kept & (claim_numeric <= 0)
        is_removable   = (is_removable | kept_no_value) & ~kept_has_value

    cleaned_df = df[~is_removable].copy()
    removed_df = df[is_removable].copy()

    s_norm  = removed_df[STATUS_COL].fillna("").astype(str).str.strip()
    s_upper = s_norm.str.upper()
    reason  = pd.Series("Status: " + s_norm, index=removed_df.index)
    reason  = reason.where(~s_upper.isin(REMOVE_STATUSES),              "Status: " + s_norm)
    reason  = reason.where(~s_upper.str.contains("CEASE", na=False),    "Status contains CEASE: " + s_norm)
    reason  = reason.where(~s_upper.str.contains("PTP",   na=False),    "PTP with no PTP Amount")
    reason  = reason.where(~s_upper.str.contains("KEPT",  na=False),    "KEPT with no Claim Paid Amount")
    reason  = reason.where(s_upper != "",                                reason)
    reason  = reason.where(~(s_upper.isin(["", "NAN"])),                "Blank Status")
    removed_df.insert(0, "Removed Reason", reason)

    srp_mask   = pd.Series([False] * len(cleaned_df), index=cleaned_df.index)
    remarks_df = pd.DataFrame()

    if REMARK_COL in cleaned_df.columns:
        remark_norm         = cleaned_df[REMARK_COL].astype(str).str.strip().str.upper()
        status_clean        = cleaned_df[STATUS_COL].astype(str).str.strip().str.upper()
        has_action_ptp      = remark_norm.str.contains(r"ACTION\s*:\s*PTP", regex=True, na=False)
        status_not_ptp_kept = (
            ~status_clean.str.contains("PTP",  na=False) &
            ~status_clean.str.contains("KEPT", na=False)
        )
        srp_mask = has_action_ptp & status_not_ptp_kept

        if srp_mask.any():
            changed_rows = cleaned_df.loc[srp_mask].copy()
            remarks_df = pd.DataFrame({
                "Row #":                    range(1, srp_mask.sum() + 1),
                STATUS_COL:                 changed_rows[STATUS_COL].values,
                REMARK_COL + " (Before)":   changed_rows[REMARK_COL].astype(str).values,
                REMARK_COL + " (After)":    changed_rows[REMARK_COL].astype(str).str.replace(
                    r"(?i)Action\s*:\s*PTP", "Action: SRP", regex=True).values,
            })

        cleaned_df.loc[srp_mask, REMARK_COL] = (
            cleaned_df.loc[srp_mask, REMARK_COL]
            .astype(str)
            .str.replace(r"(?i)Action\s*:\s*PTP", "Action: SRP", regex=True)
        )

    cols_to_drop = [
        df.columns[i]
        for i in range(COLS_TO_DROP_START, min(COLS_TO_DROP_END + 1, len(df.columns)))
    ]
    cleaned_df.drop(columns=cols_to_drop, inplace=True, errors="ignore")
    removed_df.drop(columns=cols_to_drop, inplace=True, errors="ignore")

    if DATE_COL_NAME in cleaned_df.columns:
        cleaned_df[DATE_COL_NAME] = (
            pd.to_datetime(cleaned_df[DATE_COL_NAME], errors="coerce")
            .dt.strftime("%m-%d-%Y")
        )

    if DATE_COL_NAME in cleaned_df.columns and TIME_COL_NAME in cleaned_df.columns:
        date_str = pd.to_datetime(cleaned_df[DATE_COL_NAME], errors="coerce").dt.strftime("%m-%d-%Y")
        time_str = pd.to_datetime(cleaned_df[TIME_COL_NAME], errors="coerce").dt.strftime("%I:%M:%S %p")
        cleaned_df[TIME_COL_NAME] = date_str + " " + time_str

    stats = {
        "total":       len(df),
        "retained":    len(cleaned_df),
        "removed":     len(removed_df),
        "srp_changed": int(srp_mask.sum()),
    }

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for sheet_name, frame in [("Cleaned", cleaned_df), ("Removed", removed_df)]:
            frame.to_excel(writer, index=False, sheet_name=sheet_name)
            ws = writer.sheets[sheet_name]
            normal_fill = PatternFill("solid", fgColor="1E2235")
            orange_fill = PatternFill("solid", fgColor="C05621")
            white_bold  = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.font = white_bold
                cell.fill = orange_fill if cell.value == "Removed Reason" else normal_fill
            for col in ws.columns:
                header_len = len(str(col[0].value)) if col[0].value else 10
                ws.column_dimensions[col[0].column_letter].width = min(header_len + 6, 40)

        if not remarks_df.empty:
            remarks_df.to_excel(writer, index=False, sheet_name="Remarks Changes")
            ws = writer.sheets["Remarks Changes"]
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="7C3AED")
            for col in ws.columns:
                max_len = max((len(str(c.value)) for c in col if c.value), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    buf.seek(0)
    return cleaned_df, removed_df, remarks_df, stats, buf.read()


# ── INTERACTIVE TABLE COMPONENT ───────────────────────────────────────────────
def render_interactive_table(df: pd.DataFrame, table_key: str, search_term: str = "") -> dict | None:
    """
    Renders a fully custom HTML/JS table with:
    - Auto-fit column widths (measured from content)
    - Single-cell highlight on click (only that cell, no row/col highlight)
    - Cell inspector data returned via st.session_state via a hidden input trick
    - Search term highlighting
    Returns clicked cell info or None.
    """
    if df is None or df.empty:
        st.markdown('<div class="status-bar">No rows to display.</div>', unsafe_allow_html=True)
        return None

    # Serialize dataframe to JSON (limit preview to 2000 rows for performance)
    MAX_ROWS = 2000
    display_df = df.head(MAX_ROWS).fillna("").astype(str)
    total_rows = len(df)

    columns  = list(display_df.columns)
    rows     = display_df.values.tolist()

    cols_json = json.dumps(columns)
    rows_json = json.dumps(rows)

    component_html = f"""
<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}

  body {{
    background: transparent;
    font-family: "Segoe UI", system-ui, sans-serif;
    color: #F1F5F9;
  }}

  /* ── Info bar ── */
  #info-bar {{
    font-size: 12px;
    color: #94A3B8;
    padding: 4px 0 8px;
  }}

  /* ── Inspector ── */
  #inspector {{
    display: none;
    background: #12162B;
    border: 1px solid #4F8EF7;
    border-radius: 10px;
    padding: 10px 14px;
    margin-bottom: 10px;
  }}
  #inspector.visible {{ display: block; }}
  #inspector-header {{
    display: flex;
    justify-content: space-between;
    align-items: center;
    margin-bottom: 6px;
  }}
  #inspector-col {{
    color: #4F8EF7;
    font-size: 11px;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.06em;
  }}
  #inspector-close {{
    background: none;
    border: none;
    color: #94A3B8;
    font-size: 13px;
    cursor: pointer;
    padding: 2px 6px;
    border-radius: 4px;
  }}
  #inspector-close:hover {{ color: #EF4444; background: #1E2235; }}
  #inspector-val {{
    color: #F1F5F9;
    font-size: 13px;
    line-height: 1.6;
    word-break: break-all;
    white-space: pre-wrap;
    max-height: 120px;
    overflow-y: auto;
  }}
  #inspector-val.empty {{ color: #94A3B8; font-style: italic; }}
  .search-highlight {{ background: #F59E0B; color: #000; border-radius: 2px; padding: 0 1px; }}

  /* ── Table wrapper ── */
  #table-wrap {{
    overflow-x: auto;
    overflow-y: auto;
    max-height: 520px;
    border: 1px solid #2A2D3E;
    border-radius: 10px;
  }}

  table {{
    border-collapse: collapse;
    width: max-content;
    min-width: 100%;
    table-layout: fixed;
  }}

  /* ── Header ── */
  thead tr {{
    background: #1E2235;
    position: sticky;
    top: 0;
    z-index: 10;
  }}
  th {{
    color: #94A3B8;
    font-size: 12px;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.05em;
    padding: 10px 12px;
    text-align: left;
    border-bottom: 1px solid #2A2D3E;
    border-right: 1px solid #2A2D3E;
    white-space: nowrap;
    overflow: hidden;
    text-overflow: ellipsis;
    cursor: pointer;
    user-select: none;
  }}
  th:last-child {{ border-right: none; }}
  th:hover {{ background: #252A42; color: #F1F5F9; }}
  th .sort-arrow {{ margin-left: 4px; opacity: 0.4; font-size: 10px; }}
  th.sorted-asc  .sort-arrow,
  th.sorted-desc .sort-arrow {{ opacity: 1; color: #4F8EF7; }}

  /* ── Body rows ── */
  tbody tr {{ transition: background 0.08s; }}
  tbody tr:nth-child(even) {{ background: #1A1D27; }}
  tbody tr:nth-child(odd)  {{ background: #14172A; }}
  tbody tr:hover           {{ background: #1E2747; }}

  /* ── Cells ── */
  td {{
    font-size: 13px;
    padding: 8px 12px;
    border-bottom: 1px solid #1E2235;
    border-right: 1px solid #1E2235;
    white-space: nowrap;
    overflow: hidden;
    text-overflow: ellipsis;
    cursor: pointer;
    vertical-align: middle;
  }}
  td:last-child {{ border-right: none; }}

  /* ── ONLY the clicked cell gets a highlight ── */
  td.cell-selected {{
    outline: 2px solid #4F8EF7;
    outline-offset: -2px;
    background: #1B2A4A !important;
    border-radius: 2px;
  }}

  /* ── Empty badge ── */
  .empty-badge {{
    color: #4A5568;
    font-style: italic;
    font-size: 12px;
  }}
</style>
</head>
<body>

<div id="info-bar">
  Showing <strong id="row-count">0</strong> of <strong>{total_rows}</strong> rows
  &nbsp;·&nbsp; <span id="col-count">0</span> columns
  &nbsp;·&nbsp; Click any cell to inspect full value
</div>

<div id="inspector">
  <div id="inspector-header">
    <span id="inspector-col">—</span>
    <button id="inspector-close" onclick="closeInspector()">✕ Close</button>
  </div>
  <div id="inspector-val"></div>
</div>

<div id="table-wrap">
  <table id="main-table">
    <thead id="thead"></thead>
    <tbody id="tbody"></tbody>
  </table>
</div>

<script>
const COLUMNS   = {cols_json};
const ALL_ROWS  = {rows_json};
const SEARCH    = {json.dumps(search_term.lower().strip())};

let filteredRows  = [];
let sortColIdx    = -1;
let sortAsc       = true;
let selectedCell  = null;   // reference to the currently highlighted <td>

// ── Autofit column widths ──────────────────────────────────────────────────
// Measure text width using a hidden canvas context (fast, no DOM needed)
const _canvas = document.createElement("canvas");
const _ctx    = _canvas.getContext("2d");
_ctx.font     = "13px Segoe UI, system-ui, sans-serif";

function measureText(str) {{
  return _ctx.measureText(String(str)).width;
}}

function computeColWidths(rows) {{
  const PADDING    = 28;   // cell padding left+right
  const MIN_WIDTH  = 60;
  const MAX_WIDTH  = 320;
  const HEAD_FONT  = "700 12px Segoe UI, system-ui, sans-serif";

  // Header widths
  _ctx.font    = HEAD_FONT;
  const widths = COLUMNS.map(c => Math.ceil(measureText(c.toUpperCase()) + PADDING + 20));

  // Sample up to 200 rows for body widths
  _ctx.font = "13px Segoe UI, system-ui, sans-serif";
  const sample = rows.slice(0, 200);
  sample.forEach(row => {{
    row.forEach((cell, ci) => {{
      const w = Math.ceil(measureText(String(cell)) + PADDING);
      if (w > widths[ci]) widths[ci] = w;
    }});
  }});

  return widths.map(w => Math.max(MIN_WIDTH, Math.min(MAX_WIDTH, w)));
}}

// ── Filter rows ───────────────────────────────────────────────────────────
function applyFilter() {{
  if (!SEARCH) {{
    filteredRows = [...ALL_ROWS];
  }} else {{
    filteredRows = ALL_ROWS.filter(row =>
      row.some(cell => String(cell).toLowerCase().includes(SEARCH))
    );
  }}
}}

// ── Highlight search term in a string → safe HTML ─────────────────────────
function escapeHtml(s) {{
  return String(s)
    .replace(/&/g, "&amp;")
    .replace(/</g, "&lt;")
    .replace(/>/g, "&gt;")
    .replace(/"/g, "&quot;");
}}

function highlightSearch(str) {{
  if (!SEARCH) return escapeHtml(str);
  const escaped = escapeHtml(str);
  const re      = new RegExp(SEARCH.replace(/[.*+?^${{}}()|[\\]\\\\]/g, "\\\\$&"), "gi");
  return escaped.replace(re, m => `<span class="search-highlight">${{m}}</span>`);
}}

// ── Sort ──────────────────────────────────────────────────────────────────
function sortBy(colIdx) {{
  if (sortColIdx === colIdx) {{
    sortAsc = !sortAsc;
  }} else {{
    sortColIdx = colIdx;
    sortAsc    = true;
  }}
  filteredRows.sort((a, b) => {{
    const av = String(a[colIdx]).toLowerCase();
    const bv = String(b[colIdx]).toLowerCase();
    const n  = av.localeCompare(bv, undefined, {{ numeric: true }});
    return sortAsc ? n : -n;
  }});
  renderTable();
  // Update header arrows
  document.querySelectorAll("th").forEach((th, i) => {{
    th.classList.toggle("sorted-asc",  i === sortColIdx && sortAsc);
    th.classList.toggle("sorted-desc", i === sortColIdx && !sortAsc);
    const arrow = th.querySelector(".sort-arrow");
    if (arrow) arrow.textContent = i === sortColIdx ? (sortAsc ? "▲" : "▼") : "⇅";
  }});
}}

// ── Cell click → inspector ────────────────────────────────────────────────
function onCellClick(td, colIdx, value) {{
  // Remove highlight from previously selected cell
  if (selectedCell && selectedCell !== td) {{
    selectedCell.classList.remove("cell-selected");
  }}

  // If clicking the same cell again — toggle off
  if (selectedCell === td) {{
    td.classList.remove("cell-selected");
    selectedCell = null;
    closeInspector();
    return;
  }}

  // Highlight only this cell
  td.classList.add("cell-selected");
  selectedCell = td;

  // Show inspector
  const col = COLUMNS[colIdx];
  document.getElementById("inspector-col").textContent = "📌  " + col;

  const valEl = document.getElementById("inspector-val");
  if (!value || value.trim() === "" || value === "nan") {{
    valEl.innerHTML = "(empty)";
    valEl.className = "empty";
  }} else {{
    valEl.innerHTML = highlightSearch(value);
    valEl.className = "";
  }}

  document.getElementById("inspector").classList.add("visible");
}}

function closeInspector() {{
  document.getElementById("inspector").classList.remove("visible");
  if (selectedCell) {{
    selectedCell.classList.remove("cell-selected");
    selectedCell = null;
  }}
}}

// ── Render ────────────────────────────────────────────────────────────────
function renderTable() {{
  const colWidths = computeColWidths(filteredRows);
  const thead = document.getElementById("thead");
  const tbody = document.getElementById("tbody");

  // Header
  thead.innerHTML = "";
  const hr = document.createElement("tr");
  COLUMNS.forEach((col, ci) => {{
    const th        = document.createElement("th");
    th.style.width  = colWidths[ci] + "px";
    th.innerHTML    = escapeHtml(col) + ` <span class="sort-arrow">⇅</span>`;
    th.title        = col;
    th.onclick      = () => sortBy(ci);
    if (ci === sortColIdx) {{
      th.classList.add(sortAsc ? "sorted-asc" : "sorted-desc");
      th.querySelector(".sort-arrow").textContent = sortAsc ? "▲" : "▼";
    }}
    hr.appendChild(th);
  }});
  thead.appendChild(hr);

  // Body
  tbody.innerHTML = "";
  selectedCell    = null;   // reset selection when re-rendering

  filteredRows.forEach((row, ri) => {{
    const tr = document.createElement("tr");
    row.forEach((cell, ci) => {{
      const td       = document.createElement("td");
      td.style.width = colWidths[ci] + "px";
      td.title       = String(cell);   // native tooltip on hover

      const display  = String(cell);
      if (!display || display === "nan") {{
        td.innerHTML = `<span class="empty-badge">—</span>`;
      }} else if (SEARCH && display.toLowerCase().includes(SEARCH)) {{
        td.innerHTML = highlightSearch(display);
      }} else {{
        td.textContent = display;
      }}

      td.addEventListener("click", () => onCellClick(td, ci, display));
      tr.appendChild(td);
    }});
    tbody.appendChild(tr);
  }});

  // Info bar
  document.getElementById("row-count").textContent  = filteredRows.length;
  document.getElementById("col-count").textContent  = COLUMNS.length + " columns";
}}

// ── Init ─────────────────────────────────────────────────────────────────
applyFilter();
renderTable();
</script>
</body>
</html>
"""

    height = min(700, 120 + min(len(display_df), MAX_ROWS) * 37 + 60)
    height = max(height, 260)
    components.html(component_html, height=height, scrolling=False)
    return None


# ── SESSION STATE INIT ────────────────────────────────────────────────────────
for key, default in {
    "cleaned_df":        None,
    "removed_df":        None,
    "removed_reason_df": None,
    "remarks_df":        None,
    "output_bytes":      None,
    "stats":             None,
    "processed":         False,
    "error_msg":         None,
}.items():
    if key not in st.session_state:
        st.session_state[key] = default


# ── SIDEBAR ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("""
    <div style="padding: 8px 0 20px;">
      <div style="font-size: 20px; font-weight: 800; color: #F1F5F9;">⬡ Excel Data Cleaner</div>
      <div style="font-size: 12px; color: #94A3B8; margin-top: 4px;">Drop · Preview · Verify · Download</div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown('<div class="section-header">Upload File</div>', unsafe_allow_html=True)
    uploaded_file = st.file_uploader(
        "Choose an Excel file",
        type=["xlsx", "xlsm", "xls"],
        label_visibility="collapsed",
    )

    if uploaded_file:
        st.markdown(f"""
        <div style="background:#1A1D27; border:1px solid #2A2D3E; border-radius:8px;
                    padding:8px 12px; margin: 8px 0; font-size:12px; color:#4F8EF7;">
          📄 {uploaded_file.name}
        </div>
        """, unsafe_allow_html=True)

    st.markdown('<div class="section-header">Summary</div>', unsafe_allow_html=True)

    stats = st.session_state.stats or {}
    col1, col2 = st.columns(2)
    with col1:
        st.metric("Total Rows",    stats.get("total",       "—"))
        st.metric("Rows Removed",  stats.get("removed",     "—"))
    with col2:
        st.metric("Rows Retained", stats.get("retained",    "—"))
        st.metric("Remarks Fixed", stats.get("srp_changed", "—"))

    st.markdown("<br>", unsafe_allow_html=True)

    run_btn = st.button(
        "▶  Run Cleaner",
        disabled=uploaded_file is None,
        use_container_width=True,
    )

    if st.session_state.output_bytes:
        st.download_button(
            label="⬇  Download Output",
            data=st.session_state.output_bytes,
            file_name="cleaned_output.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    st.markdown("""
    <div class="watermark">
      Created by <span>Vincent Corocoto</span> · 09567796275<br>
      <span style="color:#4A5568; font-style:italic;">
        "Kapag ang palay naging bigas, May bumayo."
      </span>
    </div>
    """, unsafe_allow_html=True)


# ── MAIN AREA ─────────────────────────────────────────────────────────────────

# Run processing
if run_btn and uploaded_file:
    with st.spinner("Processing file…"):
        try:
            file_bytes = uploaded_file.read()
            cleaned, removed, remarks, stats_out, out_bytes = process_file(file_bytes)

            # Build removed_reason_df
            if "Removed Reason" in removed.columns:
                status_cols = [c for c in removed.columns if c != "Removed Reason"]
                mask = removed["Removed Reason"].str.startswith("Status: ")
                removed_reason = removed[mask][["Removed Reason"] + status_cols].copy()
            else:
                removed_reason = removed.copy()

            st.session_state.cleaned_df        = cleaned
            st.session_state.removed_df        = removed
            st.session_state.removed_reason_df = removed_reason
            st.session_state.remarks_df        = remarks
            st.session_state.output_bytes      = out_bytes
            st.session_state.stats             = stats_out
            st.session_state.processed         = True
            st.session_state.error_msg         = None
            st.rerun()
        except Exception as e:
            st.session_state.error_msg = str(e)

if st.session_state.error_msg:
    st.error(f"Processing error: {st.session_state.error_msg}")

if not st.session_state.processed:
    # Landing state
    st.markdown("""
    <div style="display:flex; flex-direction:column; align-items:center; justify-content:center;
                height:60vh; color:#94A3B8; text-align:center;">
      <div style="font-size:56px; margin-bottom:16px;">📂</div>
      <div style="font-size:20px; font-weight:700; color:#F1F5F9; margin-bottom:8px;">
        Upload an Excel file to get started
      </div>
      <div style="font-size:14px;">
        Use the sidebar to upload a .xlsx / .xlsm / .xls file,<br>
        then click <strong style="color:#4F8EF7;">▶ Run Cleaner</strong> to process it.
      </div>
    </div>
    """, unsafe_allow_html=True)
    st.stop()

# ── TAB LAYOUT ────────────────────────────────────────────────────────────────
tabs = st.tabs([
    "✅ Cleaned Rows",
    "🗑 Removed Rows",
    "📋 Removed Status",
    "✏️ Remarks Changes",
])

tab_data = [
    (st.session_state.cleaned_df,        "cleaned"),
    (st.session_state.removed_df,        "removed"),
    (st.session_state.removed_reason_df, "removed_reason"),
    (st.session_state.remarks_df,        "remarks"),
]

empty_messages = {
    "cleaned":        "No retained rows.",
    "removed":        "No removed rows.",
    "removed_reason": "No removed rows matching exact statuses.",
    "remarks":        "No remark changes found.",
}

for tab, (df, tab_key) in zip(tabs, tab_data):
    with tab:
        if df is None or (hasattr(df, "empty") and df.empty):
            st.markdown(
                f'<div class="status-bar">{empty_messages[tab_key]}</div>',
                unsafe_allow_html=True,
            )
            continue

        # Search bar
        search_col, count_col = st.columns([3, 1])
        with search_col:
            search = st.text_input(
                "🔍 Filter rows…",
                key=f"search_{tab_key}",
                placeholder="Type to filter…",
                label_visibility="collapsed",
            )
        with count_col:
            total = len(df)
            if search:
                combined = df.fillna("").astype(str).agg(" ".join, axis=1).str.lower()
                matched  = int(combined.str.contains(search.lower().strip(), regex=False).sum())
                st.markdown(
                    f'<div style="padding-top:8px; font-size:12px; color:#94A3B8; text-align:right;">'
                    f'{matched:,} / {total:,} rows</div>',
                    unsafe_allow_html=True,
                )
            else:
                st.markdown(
                    f'<div style="padding-top:8px; font-size:12px; color:#94A3B8; text-align:right;">'
                    f'{total:,} rows · {len(df.columns)} cols</div>',
                    unsafe_allow_html=True,
                )

        render_interactive_table(df, table_key=tab_key, search_term=search or "")